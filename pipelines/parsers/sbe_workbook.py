"""Parser for the Statements of Budget Estimates workbook.

``indiabudget.gov.in/doc/eb/allsbe.xlsx`` is the Union Budget's own spreadsheet:
one sheet per Demand for Grants, each carrying four stages side by side.

    Actuals 2024-2025 | Budget Estimates 2025-2026 | Revised Estimates 2025-2026 | Budget Estimates 2026-2027

Each stage is split Revenue / Capital / Total, and each demand opens with a
Gross / Recoveries / Receipts / **Net** summary. Net is the figure that matters:
gross double counts money the demand recovers from elsewhere in government.

Two properties of this file drive the whole design.

**Column positions are not fixed.** The Total column sits at index 19 on the
first sheet and at 23 on the second, and there are several layouts across the
102 demands. Reading a fixed column would silently attribute one stage's money
to another, which is the worst failure this parser could have, so every sheet's
stage columns are located from its own header rows and a sheet whose headers
cannot be read produces a parse error rather than a guess.

**The unit is stated in the document.** "(In ₹ Crores)" appears near the title.
It is read, not assumed. A sheet that does not state its unit is not parsed,
because a bare number in a spreadsheet is exactly as ambiguous as one in a PDF.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

import structlog

from pipelines.parsers.fy_dates import fy_from_indian_label
from pipelines.parsers.text_norm import normalise_org_name

log = structlog.get_logger(__name__)

#: Header text to fiscal stage. Order matters only for readability; matching is
#: on the leading words, because the year that follows varies by edition.
STAGE_BY_HEADER: tuple[tuple[str, str], ...] = (
    ("actuals", "EXPENDITURE"),
    ("revised estimates", "RE"),
    ("budget estimates", "BE"),
)

#: The demand summary rows, in the order the statement prints them. Only Net is
#: ingested; the others are read to confirm the block was found.
SUMMARY_LABELS = ("Gross", "Recoveries", "Receipts", "Net")

#: Sheets holding a demand. `Sheet1` is the index.
SHEET_PREFIX = "sbe"

_DEMAND_NO = re.compile(r"Demand\s+No\.?\s*(\d+)", re.IGNORECASE)
_YEAR_PAIR = re.compile(r"(\d{4})\s*[-–]\s*(\d{4}|\d{2})")
_CRORE_UNIT = re.compile(r"in\s*₹?\s*crore", re.IGNORECASE)
_LAKH_UNIT = re.compile(r"in\s*₹?\s*lakh", re.IGNORECASE)


@dataclass(frozen=True, slots=True)
class StageColumn:
    """One stage of the statement and the column holding its Total."""

    stage: str
    fy: str
    total_index: int
    header: str


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _fy_from_header(header: str) -> str | None:
    """Read 'Budget Estimates 2026-2027' into our FY label."""
    match = _YEAR_PAIR.search(header)
    if not match:
        return None
    start, end = match.group(1), match.group(2)
    label = f"{start}-{end[-2:]}"
    try:
        return fy_from_indian_label(label)
    except Exception:  # noqa: BLE001 - not every digit pair is a year pair
        return None


def _stage_for(header: str) -> str | None:
    lowered = header.lower()
    for marker, stage in STAGE_BY_HEADER:
        if lowered.startswith(marker):
            return stage
    return None


def locate_stage_columns(rows: list[tuple[Any, ...]]) -> list[StageColumn]:
    """Find each stage and the column index of its Total, from the sheet itself.

    The stage header row names the stage and the year; the row beneath it splits
    that stage into Revenue, Capital and Total. The Total belonging to a stage
    is the first one at or after the stage's own column and before the next
    stage begins, which is what makes this survive the column shifts between
    sheets.
    """
    header_row_index: int | None = None
    for index, row in enumerate(rows[:20]):
        joined = " ".join(_text(value) for value in row).lower()
        if "actuals" in joined and "budget estimates" in joined:
            header_row_index = index
            break
    if header_row_index is None:
        return []

    header_row = rows[header_row_index]
    stage_starts: list[tuple[int, str]] = [
        (column, _text(value)) for column, value in enumerate(header_row) if _text(value)
    ]
    if not stage_starts:
        return []

    # The Revenue/Capital/Total row is the next row that names Total at all.
    split_row: tuple[Any, ...] | None = None
    for row in rows[header_row_index + 1 : header_row_index + 4]:
        if any(_text(value).lower() == "total" for value in row):
            split_row = row
            break
    if split_row is None:
        return []

    total_columns = [
        column for column, value in enumerate(split_row) if _text(value).lower() == "total"
    ]

    located: list[StageColumn] = []
    for position, (column, header) in enumerate(stage_starts):
        stage = _stage_for(header)
        fy = _fy_from_header(header)
        if stage is None or fy is None:
            continue
        next_column = stage_starts[position + 1][0] if position + 1 < len(stage_starts) else 10**6
        total = next((c for c in total_columns if column <= c < next_column), None)
        if total is None:
            continue
        located.append(StageColumn(stage=stage, fy=fy, total_index=total, header=header))
    return located


def _to_decimal(value: Any) -> Decimal | None:
    """Read a cell as an amount, or None when the statement reports nothing.

    The statement prints '...' for a head with no provision. That is 'not
    reported', not zero, and it must never become a number.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text or set(text) <= {".", " "}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _split_title(rows: list[tuple[Any, ...]]) -> tuple[str, str]:
    """Read the owning body and the demand's own name from the title cell.

    The title is one cell holding three lines::

        Ministry of Agriculture and Farmers Welfare
        Demand No. 1
        Department of Agriculture and Farmers Welfare

    Everything above the demand number is the body the demand belongs to, and
    that is what a ministry-level figure is attributed to. Matching on the words
    "Ministry" or "Department" instead would have been simpler and would have
    silently dropped five demands: the President, Lok Sabha, Rajya Sabha, the
    Vice-President's Secretariat and the Union Public Service Commission are
    none of those things, and they are spending public money too.
    """
    best: list[str] = []
    for row in rows[:9]:
        for value in row:
            text = _text(value).replace("\xa0", " ")
            if not _DEMAND_NO.search(text):
                continue
            lines = [" ".join(line.split()) for line in text.split("\n")]
            lines = [line for line in lines if line]
            if len(lines) > len(best):
                best = lines
    if not best:
        return "", ""

    marker = next((i for i, line in enumerate(best) if _DEMAND_NO.search(line)), None)
    if marker is None:
        return "", ""
    owner = " ".join(best[:marker]).strip(" ,-")
    name = " ".join(best[marker + 1 :]).strip(" ,-")
    # A demand whose number shares the line with the body, rather than sitting
    # on its own, still yields the body once the number is removed.
    if not owner:
        owner = _DEMAND_NO.sub("", best[marker]).strip(" ,-")
    return owner, name


def _find_summary_net(rows: list[tuple[Any, ...]], stop: int) -> tuple[Any, ...] | None:
    """The demand's own Net row, from the summary block above section A.

    Later sections repeat Net as a sub-total for a scheme group, so the search
    stops at 'A. The Budget allocation', which is where the demand summary ends
    and the detail begins.
    """
    for row in rows[:stop]:
        for value in row:
            if _text(value) == "Net":
                return row
    return None


def _section_a_index(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows):
        if any(_text(value).startswith("A. The Budget") for value in row):
            return index
    return min(len(rows), 30)


def parse_sheet(
    rows: list[tuple[Any, ...]],
    *,
    sheet: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse one demand sheet into stage rows."""
    errors: list[dict[str, Any]] = []
    header_text = " ".join(_text(value) for row in rows[:9] for value in row)

    if _LAKH_UNIT.search(header_text) and not _CRORE_UNIT.search(header_text):
        errors.append(
            {
                "reason": f"{sheet} states its amounts in lakh, which this parser does not "
                f"convert. The unit is read from the document and never assumed.",
                "stage_hint": "BE",
                "raw_context": {"sheet": sheet},
            }
        )
        return [], errors
    if not _CRORE_UNIT.search(header_text):
        errors.append(
            {
                "reason": f"{sheet} does not state its unit. A bare number in a spreadsheet is "
                f"exactly as ambiguous as one in a PDF, so nothing is read from it.",
                "stage_hint": "BE",
                "raw_context": {"sheet": sheet},
            }
        )
        return [], errors

    demand_match = _DEMAND_NO.search(header_text)
    demand_no = demand_match.group(1) if demand_match else None

    entity, demand_name = _split_title(rows)

    if not entity:
        errors.append(
            {
                "reason": f"{sheet} names no ministry or department.",
                "stage_hint": "BE",
                "raw_context": {"sheet": sheet, "demand_no": demand_no},
            }
        )
        return [], errors

    stages = locate_stage_columns(rows)
    if not stages:
        errors.append(
            {
                "reason": f"{sheet} has no readable stage header row. The statement has been "
                f"restructured and the column mapping needs revisiting.",
                "stage_hint": "BE",
                "raw_context": {"sheet": sheet, "entity": entity},
            }
        )
        return [], errors

    net = _find_summary_net(rows, _section_a_index(rows))
    if net is None:
        errors.append(
            {
                "reason": f"{sheet} has no Net row in its demand summary, so the demand total "
                f"cannot be read without adding up sections, which would double count "
                f"recoveries.",
                "stage_hint": "BE",
                "raw_context": {"sheet": sheet, "entity": entity},
            }
        )
        return [], errors

    out: list[dict[str, Any]] = []
    for stage in stages:
        amount = _to_decimal(net[stage.total_index] if stage.total_index < len(net) else None)
        if amount is None:
            # A stage with no figure is a stage the statement does not report.
            continue
        out.append(
            {
                "sheet": sheet,
                "demand_no": demand_no,
                # The body a figure is attributed to. The demand name is the
                # level the entity master records (departments, not ministries)
                # so it is preferred; the owning ministry is the fallback for a
                # demand whose name is the ministry itself.
                "entity_raw": demand_name or entity,
                "entity_normalised": normalise_org_name(demand_name or entity),
                "owner_raw": entity,
                "owner_normalised": normalise_org_name(entity),
                # Read from the publisher's own spreadsheet cell, which is a
                # different confidence from a figure recovered out of a PDF
                # table, and the provenance popover says which.
                "extraction_method": "spreadsheet",
                "fy": stage.fy,
                "stage": stage.stage,
                "amount_inr_cr": amount,
                "demand_name": demand_name,
                "header": stage.header,
            }
        )
    return out, errors


def parse_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    """Parse every demand sheet in the statements workbook.

    Returns rows, parse errors, and the stage headers seen, which the drift
    check uses as its column set: a budget that renames or reorders its stage
    columns is a document worth stopping on.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise RuntimeError("openpyxl is required to read the budget workbook.") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows_out: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    headers: set[str] = set()

    for sheet in workbook.sheetnames:
        if not sheet.lower().startswith(SHEET_PREFIX):
            continue
        rows = list(workbook[sheet].iter_rows(values_only=True))
        parsed, sheet_errors = parse_sheet(rows, sheet=sheet)
        rows_out.extend(parsed)
        errors.extend(sheet_errors)
        headers.update(row["header"] for row in parsed)

    log.info(
        "sbe.parsed",
        sheets=len(workbook.sheetnames),
        rows=len(rows_out),
        errors=len(errors),
    )
    return rows_out, errors, sorted(headers)
